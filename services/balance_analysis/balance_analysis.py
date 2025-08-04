"""Balance Analysis Automation Service"""

import io
import re
import unicodedata
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from typing import List, Union, Tuple


def normalize_text(text):
    """Normalizes text by removing accents and converting to lowercase."""
    if not text:
        return ''
    nfkd = unicodedata.normalize('NFD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower()


def _get_pdf_obj(pdf_source):
    """
    Retorna um objeto file-like para pdfplumber.open,
    aceitando bytes ou file-like.
    """
    if isinstance(pdf_source, (bytes, bytearray)):
        return io.BytesIO(pdf_source)
    return pdf_source


def extract_entity_name(pdf_source):
    """Extracts the entity name from the PDF, looking for 'Entidade:'."""
    pattern = r"Entidade:\s*(.+)"
    with pdfplumber.open(_get_pdf_obj(pdf_source)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            m = re.search(pattern, text)
            if m:
                return m.group(1).strip()
    raise ValueError("Não foi possível encontrar 'Entidade:' no PDF.")


def extract_section_types(pdf_source):
    """Extracts section types from the PDF to determine if it contains 'balanco' or 'dre'."""
    keywords = {
        'balanco': 'balanco patrimonial',
        'dre': 'demonstracao de resultado do exercicio'
    }
    found = set()
    with pdfplumber.open(_get_pdf_obj(pdf_source)) as pdf:
        for page in pdf.pages:
            text = normalize_text(page.extract_text() or '')
            for key, phrase in keywords.items():
                if phrase in text:
                    found.add(key)
    return found


def extract_final_balance_by_label(pdf_source: str, label_text: str) -> str:
    """Extracts the final balance value from a PDF based on a specified label text."""
    label_norm = normalize_text(label_text)
    with pdfplumber.open(_get_pdf_obj(pdf_source)) as pdf:
        for page in pdf.pages:
            tables = page.find_tables()
            for table in tables:
                rows = [[cell.strip() if cell else '' for cell in r] for r in table.extract()]
                header = rows[0]
                low = [normalize_text(h) for h in header]
                if 'saldo final' in low:
                    idx = low.index('saldo final')
                    for data_row in rows[1:]:
                        for cell in data_row:
                            if normalize_text(cell) == label_norm:
                                val = data_row[idx]
                                if val:
                                    return val
            text = page.extract_text() or ''
            text_norm = normalize_text(text)
            pattern = rf"{re.escape(label_norm)}[^\d]*?([\d\.,()]+)[^\d]*?([\d\.,()]+)"
            m = re.search(pattern, text_norm, re.IGNORECASE)
            if m:
                return m.group(2)
    raise ValueError(f"Não foi possível encontrar '{label_text}' para Saldo Final no PDF.")


def update_balance_sheet(
        excel_path: str,
        value: float,
        cell: str,
        sheet_name: str = None,
        is_currency=True
    ) -> None:
    """Updates the specified cell in an Excel sheet with the given value."""

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    ws[cell] = value
    if is_currency:
        ws[cell].number_format = '"R$" #,##0.00'
    else:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(excel_path)


def parse_currency_str(val_str: str) -> float:
    """Parses a Brazilian formatted currency string into a float."""

    s = val_str.strip()
    negative = False
    if s.startswith('(') and s.endswith(')'):
        negative = True
        s = s[1:-1].strip()
    clean = s.replace('.', '').replace(',', '.')
    try:
        num = float(clean)
    except ValueError:
        raise ValueError(f"Não foi possível converter '{val_str}' em número.")
    return -num if negative else num


def handle_balanco(pdf_source, excel_path, column_prefix, sheet_name=None):
    """Handles the balance sheet extraction and updates the Excel file."""

    entity = extract_entity_name(pdf_source)
    update_balance_sheet(excel_path, entity, "B3", sheet_name, is_currency=False)

    mapping = {
        "ATIVO": 7,
        "DISPONÍVEL": 8,
        "ATIVO CIRCULANTE": 9,
        "CONTAS A RECEBER": 10,
        "ESTOQUES": 11,
        "IMOBILIZADO": 12,
        "ATIVO NÃO CIRCULANTE": 13,
        "PASSIVO": 14,
        "PASSIVO CIRCULANTE": 15,
        "FORNECEDORES": 16,
        "SALARIOS E ENCARGOS": 17,
        "TRIBUTOS A RECOLHER": 18,
        "EMPRÉSTIMOS E FINANCIAMENTOS": 19,
        "PASSIVO NÃO CIRCULANTE": 20,
        "PATRIMONIO LIQUIDO": 21,
    }
    for label, row in mapping.items():
        try:
            # Reaproveitar função de extração de saldo final
            val_str = extract_final_balance_by_label(pdf_source, label)
            val_num = parse_currency_str(val_str)
            cell = f"{column_prefix}{row}"
            update_balance_sheet(excel_path, val_num, cell, sheet_name, is_currency=True)
            print(f"[Balanço] '{label}' -> {cell} = {val_num}")
        except Exception as e:
            print(f"[Balanço] erro ao processar '{label}': {e}")


def handle_dre(pdf_source, excel_path, column_prefix, sheet_name=None):
    """
    Lógica inicial para preencher a aba de Demonstração de Resultado do Exercício.
    Implementar mapeamento e extração específicos.
    """
    dre_mapping: List[Tuple[Union[str, List[str]], int]]= [
        (["RECEITA OPERACIONAL", "RECEITA LIQUIDA"], 8),
        (["RECEITA OPERACIONAL", "RECEITA LIQUIDA"], 9),
        (["CUSTOS OPERACIONAIS", "CUSTO DAS VENDAS/SERVICOS"], 10),
        ("DESPESAS OPERACIONAIS", 11),
        ("DESPESAS FINANCEIRAS", 14),
        (["OUTRAS DESPESAS/RECEITAS", "OUTRAS RECEITAS E DESPESAS"], 15),
        (["LUCRO (PREJUIZO) LIQUIDO DO EXERCICIO", "RESULTADO LIQUIDO"], 17),
    ]

    for labels, row in dre_mapping:
        found = False
        if isinstance(labels, list):
            for lbl in labels:
                try:
                    val_str = extract_final_balance_by_label(pdf_source=pdf_source, label_text=lbl)
                    found = True
                    break
                except ValueError:
                    continue
            if not found:
                print(
                    f"[DRE] não foi possível encontrar nenhum dos rótulos {labels} na linha {row}"
                )
                continue
        else:
            try:
                val_str = extract_final_balance_by_label(
                    pdf_source=pdf_source,
                    label_text=labels
                )
                found = True
            except ValueError as e:
                print(f"[DRE] erro ao processar '{labels}' na linha {row}: {e}")
                continue
        if found:
            try:
                val_num = parse_currency_str(val_str)
                cell = f"{column_prefix}{row}"
                update_balance_sheet(
                    excel_path=excel_path,
                    value=val_num,
                    cell=cell,
                    sheet_name=sheet_name,
                    is_currency=True
                )
            except Exception as e:
                print(f"[DRE] erro ao converter valor na linha {row}: {e}")


def process_balance_analysis_pdf(
    pdf_bytes,
    balanco_column_prefix,
    dre_column_prefix
):
    """Main function to update balance analysis from PDF to Excel."""

    sections = extract_section_types(pdf_source=pdf_bytes)
    if 'balanco' in sections:
        handle_balanco(
            pdf_source=pdf_bytes,
            excel_path='static/files/analise_balanco_modelo.xlsx',
            column_prefix=balanco_column_prefix,
            sheet_name='COMPARATIVO BALANÇO'
        )
    if 'dre' in sections:
        handle_dre(
            pdf_source=pdf_bytes,
            excel_path='static/files/analise_balanco_modelo.xlsx',
            column_prefix=dre_column_prefix,
            sheet_name='DRE e CICLO'
        )
    if not sections:
        print('Seção não identificada no PDF. Aguarde implementação.')
