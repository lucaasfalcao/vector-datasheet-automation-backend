"""Balance Analysis Automation Service"""

import re
import pdfplumber
from openpyxl import load_workbook


def extract_saldo_final_by_label(pdf_path, label_text):
    '''
    Abre o PDF e extrai o valor da coluna 'Saldo Final' na linha que contém o texto especificado.
    '''
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Tenta extração de tabelas
            tables = page.find_tables()
            for table in tables:
                rows = [[cell.strip() if cell else '' for cell in r] for r in table.extract()]
                header = rows[0]
                low = [h.lower() for h in header]
                if 'saldo final' in low:
                    idx = low.index('saldo final')
                    for data_row in rows[1:]:
                        # verifica se alguma célula emparelha com o label
                        if any((cell or '').lower() == label_text.lower() for cell in data_row):
                            val = data_row[idx]
                            if val:
                                return val
            # Fallback por regex no texto
            text = page.extract_text() or ''
            pattern = rf"{re.escape(label_text)}\D*([\d\.\,]+)"
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                return m.group(1)
    raise ValueError(f"Não foi possível encontrar '{label_text}' para Saldo Final no PDF.")


# Função para atualizar a célula B7 na planilha Excel
def update_balance_sheet(excel_path, value, cell, sheet_name=None):
    """Updates the specified cell in an Excel sheet with the given value."""
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    # Define valor numérico
    ws[cell] = value
    # Formatação de moeda: R$ com separadores de milhar e duas casas decimais
    ws[cell].number_format = '"R$" #,##0.00'
    wb.save(excel_path)


def parse_currency_str(val_str):
    '''
    Converte string de formato brasileiro (p.ex. '1.234.567,89') em float.
    '''
    clean = val_str.replace('.', '').replace(',', '.')
    return float(clean)


mapping = {
        "ATIVO": "B7",
        "DISPONÍVEL": "B8",
        "ATIVO CIRCULANTE": "B9",
        "CONTAS A RECEBER": "B10",
        "ESTOQUES": "B11",
        "IMOBILIZADO": "B12",
        "ATIVO NÃO CIRCULANTE": "B13",
        "PASSIVO": "B14",
        "PASSIVO CIRCULANTE": "B15",
        "FORNECEDORES": "B16",
        "SALARIOS E ENCARGOS S/FOLHA DE": "B17",
        "TRIBUTOS A RECOLHER": "B18",
        "EMPRESTIMOS E FINANCIAMENTOS": "B19",
        "PASSIVO NÃO CIRCULANTE": "B20",
        "PATRIMONIO LIQUIDO": "B21",
    }

for label, cell in mapping.items():
    try:
        val_str = extract_saldo_final_by_label(pdf_path='static/files/balanco_2021.pdf', label_text=label)
        val_num = parse_currency_str(val_str)
        update_balance_sheet(excel_path='static/files/analise_balanco_modelo.xlsx', value=val_num, cell=cell)
        print(f"Valor {val_num} inserido em {cell} para '{label}'.")
    except ValueError as e:
        print(e)

# AJUSTAR PARA VALORES DE SALDO FINAL (ESTÁ PEGANDO O INICIAL)!!!